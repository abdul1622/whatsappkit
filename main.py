from core import create_app
from flask import jsonify,request
import openpyxl
from datetime import date,datetime
from flask_apscheduler import APScheduler
import time
from pywhatkit.whats import sendwhats_image

app = create_app()
app.debug = True
scheduler = APScheduler()
scheduler.init_app(app)


if __name__ == "__main__":
    app.run

def hello():
    print(hello)

@app.route('/index',methods=['GET'])
def index():
    folder = 'E:\Programming'
    group = 'Project batch 7'
    xml_file = (openpyxl.load_workbook("details.xlsx")).active
    today = date.today()
    persons = []
    birth_day_persens = []
    keys =[]
    h = list(xml_file.rows)[0]
    for i in h:
        keys.append(i.value)
    print(h,keys)
    for row in xml_file.iter_rows(2, xml_file.max_row):
        data = {}
        for i in range(0,len(row)):
            data[keys[i]] = row[i].value
        persons.append(data)
                
    print(persons)
    for i in persons:
        if isinstance(i['dob'],str):
            print(i['dob'])
            i['dob'] = datetime.strptime(i['dob'],'%d/%m/%Y')
        if i['dob'].day == today.day and i['dob'].month == today.month:
            birth_day_persens.append(i)

    print(birth_day_persens ,'hi')
    for i in birth_day_persens:
        image_path = folder+'/'+ i['image'] +'.png'
        # sendwhats_image(group,image_path,'hi')
    return jsonify({'data':'success'})
# import time


def print_date_time():
    print(time.strftime("%A, %d. %B %Y %I:%M:%S %p"))

# to start sheduler path to xl file and time message
@app.route('/start',methods=['POST'])
def start():
    data = request.json
    if data['job'] == 'stop'  and scheduler.state:
        scheduler.shutdown()
        return data
    file_path = data['path']
    hour = data['time'][:2]
    minute = data['time'][3:5]
    app.apscheduler.add_job(func=print_date_time, trigger='cron',day_of_week='mon-sun',hour=hour,minute=minute,id='dob message')  
    if not scheduler.state:
        scheduler.start()
    return data
    

@app.route('/state',methods=['GET'])
def state():
    print(scheduler.state)
    return {'data':scheduler.state}
# scheduler = BackgroundScheduler()
# scheduler.add_job(func=print_date_time, trigger="interval", seconds=10)
# app.apscheduler.add_job(func=print_date_time, trigger='interval', seconds=10,id='id')


# Shut down the scheduler when exiting the app
# atexit.register(lambda: scheduler.shutdown())