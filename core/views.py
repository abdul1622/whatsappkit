from flask import Blueprint
from flask import jsonify,request
import openpyxl
from datetime import date,datetime
from flask_apscheduler import APScheduler
import time
from pywhatkit.whats import sendwhats_image,sendwhatmsg_instantly
from . import scheduler

views = Blueprint('views',__name__)

def task(*args):
    persons = args[0]
    image_folder = args[1]
    group_id = args[2]
    birth_day_persons = []
    today = date.today()
    for i in persons:
        if isinstance(i['dob'],str):
            print(i['dob'])
            i['dob'] = datetime.strptime(i['dob'],'%d/%m/%Y')
        if i['dob'].day == today.day and i['dob'].month == today.month:
            birth_day_persons.append(i)

    print(birth_day_persons ,'hi')
    for i in birth_day_persons:
        message = i['caption']
        image_path = "{}/{}".format(image_folder,i['image'])
        print(image_path)
        sendwhats_image(group_id,image_path,message,60,20)


# to start sheduler path to xl file and time message
@views.route('/start',methods=['POST'])
def schedule_task():
    data = request.json
    file_path = data['path']
    folder = data['image folder']
    group_id = data['group id']
    hour = data['time'][:2]
    minute = data['time'][3:5]
    if data['job'] == 'stop'  and scheduler.state:
        scheduler.shutdown()
        return data

    xml_file = (openpyxl.load_workbook(file_path)).active
    first_row = list(xml_file.rows)[0]
    keys =[]
    for i in first_row:
        keys.append(i.value)
    persons = []
    for row in xml_file.iter_rows(2, xml_file.max_row):
        xml_data = {}
        for i in range(0,len(row)):
            xml_data[keys[i]] = row[i].value
        persons.append(xml_data)

    scheduler.remove_all_jobs()

    scheduler.add_job(func=task,args=[persons,folder,group_id], trigger='cron',day_of_week='mon-sun',hour=hour,minute=minute,id='dob message')
    h = scheduler.get_job('dob message')  
    print(h)
    if not scheduler.state:
        scheduler.start()
    return {'status':'success'}
